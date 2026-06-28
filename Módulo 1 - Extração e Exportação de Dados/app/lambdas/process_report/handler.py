"""
Lambda: process_report
Consome mensagens do SQS, busca dados do RDS/Athena e gera relatório CSV/XLSX/PDF no S3.
Em seguida invoca notify_user via S3 Event (ou chamada direta).
"""
import json
import os
import io
import csv
import boto3
import logging
from datetime import datetime, timezone

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3      = boto3.client("s3")
lambda_ = boto3.client("lambda")

BUCKET          = os.environ["REPORTS_BUCKET"]
NOTIFY_FUNCTION = os.environ.get("NOTIFY_FUNCTION_NAME", "dna-automacao-notify-user")


def lambda_handler(event, context):
    for record in event["Records"]:
        try:
            message = json.loads(record["body"])
            _process_job(message)
        except Exception as exc:
            logger.error(f"Erro ao processar job: {exc}", exc_info=True)
            raise  # Devolve à fila para reprocessamento / DLQ


def _process_job(msg: dict):
    job_id      = msg["job_id"]
    fmt         = msg["format"]
    report_type = msg["report_type"]
    date_from   = msg.get("date_from")
    date_to     = msg.get("date_to")
    abac_scope  = msg.get("abac_scope", {})

    logger.info(f"Processando job_id={job_id} formato={fmt}")

    # 1. Buscar dados (RDS para hot data, Athena para cold)
    rows, columns = _fetch_data(report_type, date_from, date_to, abac_scope)

    # 2. Gerar arquivo no formato solicitado
    file_buffer, content_type, extension = _generate_file(rows, columns, fmt)

    # 3. Upload para S3
    s3_key = f"reports/{datetime.now(timezone.utc).strftime('%Y/%m/%d')}/{job_id}.{extension}"
    s3.put_object(
        Bucket      = BUCKET,
        Key         = s3_key,
        Body        = file_buffer.getvalue(),
        ContentType = content_type,
        Metadata    = {
            "job_id":     job_id,
            "user_email": msg.get("user_email", ""),
            "report_type": report_type,
        },
    )
    logger.info(f"Arquivo salvo em s3://{BUCKET}/{s3_key}")

    # 4. Invocar notify_user
    notify_payload = {
        "job_id":     job_id,
        "s3_key":     s3_key,
        "user_email": msg["user_email"],
        "format":     fmt,
        "report_type": report_type,
    }
    lambda_.invoke(
        FunctionName    = NOTIFY_FUNCTION,
        InvocationType  = "Event",  # assíncrono
        Payload         = json.dumps(notify_payload).encode(),
    )


def _fetch_data(report_type: str, date_from, date_to, abac_scope: dict):
    """
    Placeholder: substituir pela query real ao RDS (psycopg2) ou Athena (boto3).
    Retorna (rows: list[dict], columns: list[str])
    """
    columns = ["id", "timestamp", "device", "status", "value", "plant", "department"]
    rows = [
        {"id": 1, "timestamp": "2025-01-01T08:00:00Z", "device": "PLC-01", "status": "OK",    "value": 98.5, "plant": abac_scope.get("plant", "SP"), "department": abac_scope.get("department", "ENG")},
        {"id": 2, "timestamp": "2025-01-01T08:05:00Z", "device": "PLC-02", "status": "ALERT", "value": 45.2, "plant": abac_scope.get("plant", "SP"), "department": abac_scope.get("department", "ENG")},
    ]
    return rows, columns


def _generate_file(rows, columns, fmt: str):
    if fmt == "csv":
        return _to_csv(rows, columns)
    elif fmt == "xlsx":
        return _to_xlsx(rows, columns)
    elif fmt == "pdf":
        return _to_pdf(rows, columns)
    else:
        raise ValueError(f"Formato não suportado: {fmt}")


def _to_csv(rows, columns):
    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(wrapper, fieldnames=columns)
    writer.writeheader()
    writer.writerows(rows)
    wrapper.flush()
    buf.seek(0)
    return buf, "text/csv", "csv"


def _to_xlsx(rows, columns):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


def _to_pdf(rows, columns):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Relatório DNA Automação", styles["Title"]))

    table_data = [columns] + [[str(row.get(c, "")) for c in columns] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf, "application/pdf", "pdf"
